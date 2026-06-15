
#!/usr/bin/env python3
"""
懂车帝汽车销量榜 Excel 生成器
从懂车帝 API 获取车型销量数据，聚合品牌销量，生成带筛选功能的 Excel。
"""
import argparse
import json
import sys
import os
from datetime import datetime

try:
    import requests
except ImportError:
    os.system(f"{sys.executable} -m pip install requests -q")
    import requests

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter


def fetch_sales_data(month: str, count: int = 100) -> list:
    """从懂车帝 API 获取车型销量数据"""
    url = "https://www.dongchedi.com/motor/pc/car/rank_data"
    params = {
        "month": month,
        "count": str(count),
        "rank_data_type": "11"
    }
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
        "Referer": "https://www.dongchedi.com/sales"
    }
    
    all_items = []
    page = 0
    while True:
        params["offset"] = str(page * count)
        resp = requests.get(url, params=params, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        
        data_block = data.get("data")
        if not data_block:
            raise Exception(f"API 返回无数据: {data.get('message', '未知错误')}")
        
        items = data_block.get("list", [])
        if not items:
            break
        
        all_items.extend(items)
        
        # 检查是否还有更多页
        paging = data_block.get("paging", {})
        if not paging.get("has_more"):
            break
        page += 1
    
    return all_items


def aggregate_brand_sales(car_items: list) -> list:
    """从车型数据聚合品牌销量"""
    brand_map = {}
    for item in car_items:
        brand_name = item.get("brand_name", "")
        if not brand_name:
            continue
        count = item.get("count", 0)
        min_price = item.get("min_price", 0)
        max_price = item.get("max_price", 0)
        
        if brand_name not in brand_map:
            brand_map[brand_name] = {
                "brand_name": brand_name,
                "total_count": 0,
                "model_count": 0,
                "min_price_sum": 0,
                "max_price_sum": 0,
            }
        
        brand_map[brand_name]["total_count"] += count
        brand_map[brand_name]["model_count"] += 1
        brand_map[brand_name]["min_price_sum"] += min_price
        brand_map[brand_name]["max_price_sum"] += max_price
    
    brand_list = list(brand_map.values())
    brand_list.sort(key=lambda x: x["total_count"], reverse=True)
    
    # 添加排名
    for i, b in enumerate(brand_list):
        b["rank"] = i + 1
    
    return brand_list


def style_header(ws, headers, fill_color="1F4E79", font_color="FFFFFF"):
    """设置表头样式"""
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color=font_color)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def style_data_rows(ws, start_row, end_row, num_cols):
    """设置数据行样式"""
    data_font = Font(name="微软雅黑", size=10)
    data_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            if (row_idx - start_row) % 2 == 1:
                cell.fill = alt_fill


def write_car_sheet(wb, car_items: list, month_label: str):
    """写入车型销量榜 Sheet"""
    ws = wb.active
    ws.title = "车型销量榜"
    
    headers = ["排名", "排名变化", "车型名称", "品牌", "指导价（万）", "销量（辆）", "点评数"]
    style_header(ws, headers)
    
    for i, item in enumerate(car_items):
        row = i + 2
        rank = item.get("rank", i + 1)
        last_rank = item.get("last_rank", 0)
        rank_change = last_rank - rank if last_rank else 0
        
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=rank_change)
        ws.cell(row=row, column=3, value=item.get("series_name", ""))
        ws.cell(row=row, column=4, value=item.get("brand_name", ""))
        ws.cell(row=row, column=5, value=f"{item.get('min_price',0)}-{item.get('max_price',0)}")
        ws.cell(row=row, column=6, value=item.get("count", 0))
        ws.cell(row=row, column=7, value=item.get("car_review_count", 0))
    
    end_row = len(car_items) + 1
    style_data_rows(ws, 2, end_row, len(headers))
    
    # 条件格式：排名变化
    from openpyxl.formatting.rule import CellIsRule
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    
    change_col = get_column_letter(2)
    ws.conditional_formatting.add(
        f"{change_col}2:{change_col}{end_row}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill, font=green_font)
    )
    ws.conditional_formatting.add(
        f"{change_col}2:{change_col}{end_row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill, font=red_font)
    )
    
    # 添加筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{end_row}"
    
    # 设置列宽
    col_widths = [6, 10, 22, 16, 14, 12, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 销量列数字格式
    for row in range(2, end_row + 1):
        ws.cell(row=row, column=6).number_format = '#,##0'


def write_brand_sheet(wb, brand_items: list, month_label: str):
    """写入品牌销量榜 Sheet"""
    ws = wb.create_sheet("品牌销量榜")
    
    headers = ["排名", "品牌名称", "总销量（辆）", "旗下车型数", "均价区间（万）", "市场份额"]
    style_header(ws, headers, fill_color="2E75B6")
    
    total_all = sum(b["total_count"] for b in brand_items)
    
    for i, b in enumerate(brand_items):
        row = i + 2
        avg_min = round(b["min_price_sum"] / b["model_count"], 1) if b["model_count"] else 0
        avg_max = round(b["max_price_sum"] / b["model_count"], 1) if b["model_count"] else 0
        share = round(b["total_count"] / total_all * 100, 2) if total_all else 0
        
        ws.cell(row=row, column=1, value=b["rank"])
        ws.cell(row=row, column=2, value=b["brand_name"])
        ws.cell(row=row, column=3, value=b["total_count"])
        ws.cell(row=row, column=4, value=b["model_count"])
        ws.cell(row=row, column=5, value=f"{avg_min}-{avg_max}")
        ws.cell(row=row, column=6, value=f"{share}%")
    
    end_row = len(brand_items) + 1
    style_data_rows(ws, 2, end_row, len(headers))
    
    # 添加筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{end_row}"
    
    # 设置列宽
    col_widths = [6, 18, 16, 16, 16, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 销量列数字格式
    for row in range(2, end_row + 1):
        ws.cell(row=row, column=3).number_format = '#,##0'
    
    # 写入数据说明行
    info_row = end_row + 2
    ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=6)
    info_cell = ws.cell(row=info_row, column=1, 
                        value=f"数据来源：懂车帝 | 统计月份：{month_label} | 单位：辆 | 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    info_cell.font = Font(name="微软雅黑", size=9, color="888888", italic=True)
    info_cell.alignment = Alignment(horizontal="left", vertical="center")


def main():
    parser = argparse.ArgumentParser(description="懂车帝汽车销量榜 Excel 生成器")
    parser.add_argument("--month", "-m", required=True, help="月份，格式 YYYYMM，如 202605")
    parser.add_argument("--count", "-c", type=int, default=100, help="获取车型数量（默认100）")
    parser.add_argument("--output", "-o", required=True, help="输出 Excel 文件路径")
    parser.add_argument("--json-input", "-j", help="可选：本地 JSON 数据文件路径（跳过 API 请求）")
    args = parser.parse_args()
    
    try:
        if args.json_input:
            print(f"从本地文件加载数据: {args.json_input}")
            with open(args.json_input, "r", encoding="utf-8") as f:
                car_items = json.load(f)
        else:
            print(f"正在从懂车帝获取 {args.month} 销量数据...")
            car_items = fetch_sales_data(args.month, args.count)
        
        if not car_items:
            print("错误：未获取到任何数据")
            sys.exit(1)
        
        print(f"已获取 {len(car_items)} 条车型数据")
        
        # 聚合品牌销量
        print("正在聚合品牌销量...")
        brand_items = aggregate_brand_sales(car_items)
        print(f"已聚合 {len(brand_items)} 个品牌")
        
        # 格式化月份标签
        month_label = f"{args.month[:4]}年{args.month[4:]}月"
        
        # 生成 Excel
        print(f"正在生成 Excel: {args.output}")
        wb = Workbook()
        write_car_sheet(wb, car_items, month_label)
        write_brand_sheet(wb, brand_items, month_label)
        
        # 确保输出目录存在
        os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)
        wb.save(args.output)
        print(f"✅ 完成！Excel 已保存至: {args.output}")
        print(f"   - Sheet 1: 车型销量榜 ({len(car_items)} 款车型)")
        print(f"   - Sheet 2: 品牌销量榜 ({len(brand_items)} 个品牌)")
        
    except Exception as e:
        print(f"❌ 错误: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
